#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para migrar dados da planilha Excel para o banco de dados SQLite.
Execute UMA VEZ antes de usar a aplicação.

Uso:
    python migrate_excel_to_db.py --xlsx PLANILHA_ESTRATEGICA_MARCOS_2026.xlsx --db marcos.db

Ou com defaults:
    python migrate_excel_to_db.py
"""

import os
import sys
from datetime import datetime, date
from openpyxl import load_workbook
import json
import argparse

# Configurar o app Flask
def setup_app(db_path):
    from flask import Flask
    from models import db, Despesa, Recebimento, Categoria, Cliente, Caixa
    
    app = Flask(__name__)
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    db.init_app(app)
    
    return app, db, Despesa, Recebimento, Categoria, Cliente, Caixa


def to_date(val):
    """Converter valor para date"""
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except Exception:
        return None


def calc_status(venc, liq):
    """Calcular status da despesa"""
    if str(liq).strip() == 'Sim':
        return 'Pago'
    return 'Atrasado' if venc < date.today() else 'Pendente'


def calc_prio(cat, liq, dias, status):
    """Calcular prioridade"""
    if str(liq).strip() == 'Sim':
        return 'Pago'
    if status == 'Atrasado' or cat in ('IMPOSTO', 'FOLHA DE PAGAMENTO'):
        return '1-CRITICO'
    if dias <= 7 or cat in ('ABASTECIMENTO', 'LUZ E AGUA'):
        return '2-ALTO'
    if dias <= 30:
        return '3-MEDIO'
    return '4-BAIXO'


def migrate(xlsx_path, db_path, verbose=False):
    """Executar migração"""
    
    if not os.path.exists(xlsx_path):
        print(f"❌ Erro: Arquivo {xlsx_path} não encontrado")
        return False
    
    print(f"📂 Lendo planilha: {xlsx_path}")
    print(f"💾 Banco de dados: {db_path}")
    
    # Setup Flask app
    app, db, Despesa, Recebimento, Categoria, Cliente, Caixa = setup_app(db_path)
    
    with app.app_context():
        # Criar tabelas
        print("\n🔨 Criando tabelas...")
        db.create_all()
        
        # Limpar dados antigos (opcional)
        # Despesa.query.delete()
        # Recebimento.query.delete()
        # Cliente.query.delete()
        # Categoria.query.delete()
        # db.session.commit()
        
        # ─── Carregar planilha
        wb = load_workbook(xlsx_path, data_only=True)
        
        # ─── 1. DESPESAS
        print("\n📥 Importando DESPESAS...")
        ws_desp = wb['DADOS']
        count_desp = 0
        for r in range(2, ws_desp.max_row + 1):
            venc = to_date(ws_desp.cell(r, 1).value)
            val = ws_desp.cell(r, 7).value
            
            if not venc or val is None:
                continue
            
            # Verificar se já existe
            existing = Despesa.query.filter_by(vencimento=venc, valor_previsto=float(val)).first()
            if existing:
                if verbose:
                    print(f"  ⏭️  Despesa duplicada em linha {r}, ignorando")
                continue
            
            liq = str(ws_desp.cell(r, 9).value or 'Não').strip()
            cat = str(ws_desp.cell(r, 4).value or '').strip()
            dias = (venc - date.today()).days
            st = calc_status(venc, liq)
            prio = calc_prio(cat, liq, dias, st)
            
            desp = Despesa(
                vencimento=venc,
                mes=str(ws_desp.cell(r, 2).value or '').strip(),
                fornecedor=str(ws_desp.cell(r, 3).value or '').strip(),
                categoria=cat,
                descricao=str(ws_desp.cell(r, 5).value or '').strip(),
                forma_pgto=str(ws_desp.cell(r, 6).value or '').strip(),
                valor_previsto=float(val),
                valor_pago=float(ws_desp.cell(r, 8).value or 0),
                liquidado=liq,
                status=st,
                dias=dias,
                prioridade=prio,
            )
            db.session.add(desp)
            count_desp += 1
        
        db.session.commit()
        print(f"  ✅ {count_desp} despesas importadas")
        
        # ─── 2. RECEBIMENTOS
        print("\n📥 Importando RECEBIMENTOS...")
        ws_recb = wb['CONTAS A RECEBER']
        
        # Encontrar linha do TOTAL
        lim = ws_recb.max_row + 1
        for r in range(5, ws_recb.max_row + 1):
            if ws_recb.cell(r, 3).value in ('TOTAL:', 'A RECEBER (pendente):'):
                lim = r
                break
        
        count_recb = 0
        for r in range(5, lim):
            venc = to_date(ws_recb.cell(r, 1).value)
            esp = ws_recb.cell(r, 4).value
            
            if not venc or not esp:
                continue
            
            esp = float(esp)
            rec = float(ws_recb.cell(r, 5).value or 0)
            
            # Verificar duplicata
            existing = Recebimento.query.filter_by(vencimento=venc, valor_esperado=esp).first()
            if existing:
                if verbose:
                    print(f"  ⏭️  Recebimento duplicado em linha {r}, ignorando")
                continue
            
            if rec >= esp and esp > 0:
                st = 'Recebido'
            elif venc < date.today():
                st = 'Atrasado'
            else:
                st = 'A Receber'
            
            recb = Recebimento(
                vencimento=venc,
                cliente=str(ws_recb.cell(r, 2).value or '').strip(),
                descricao=str(ws_recb.cell(r, 3).value or '').strip(),
                valor_esperado=esp,
                valor_recebido=rec,
                status=st,
            )
            db.session.add(recb)
            count_recb += 1
        
        db.session.commit()
        print(f"  ✅ {count_recb} recebimentos importados")
        
        # ─── 3. CAIXA
        print("\n📥 Importando CAIXA...")
        try:
            saldo = float(wb['PAINEL DE CONTROLE']['B5'].value or 0)
            caixa = Caixa.query.first()
            if not caixa:
                caixa = Caixa(id=1, saldo=saldo)
                db.session.add(caixa)
            else:
                caixa.saldo = saldo
            db.session.commit()
            print(f"  ✅ Saldo de caixa: R$ {saldo:,.2f}")
        except Exception as e:
            print(f"  ⚠️  Erro ao ler caixa: {e}")
        
        # ─── 4. CATEGORIAS (de categorias.json)
        print("\n📥 Importando CATEGORIAS...")
        if os.path.exists('categorias.json'):
            try:
                with open('categorias.json', 'r', encoding='utf-8') as f:
                    cats_json = json.load(f)
                
                count_cats = 0
                for cat_name in cats_json.get('categorias_pagar', []):
                    existing = Categoria.query.filter_by(tipo='pagar', nome=cat_name).first()
                    if not existing:
                        cat = Categoria(tipo='pagar', nome=cat_name)
                        db.session.add(cat)
                        count_cats += 1
                
                for cat_name in cats_json.get('categorias_receber', []):
                    existing = Categoria.query.filter_by(tipo='receber', nome=cat_name).first()
                    if not existing:
                        cat = Categoria(tipo='receber', nome=cat_name)
                        db.session.add(cat)
                        count_cats += 1
                
                for cliente_name in cats_json.get('clientes', []):
                    existing = Cliente.query.filter_by(nome=cliente_name).first()
                    if not existing:
                        cliente = Cliente(nome=cliente_name)
                        db.session.add(cliente)
                        count_cats += 1
                
                db.session.commit()
                print(f"  ✅ {count_cats} categorias e clientes importados")
            except Exception as e:
                print(f"  ⚠️  Erro ao ler categorias.json: {e}")
        else:
            print("  ⚠️  categorias.json não encontrado, pulando")
        
        print("\n✅ MIGRAÇÃO CONCLUÍDA COM SUCESSO!")
        print(f"\n📊 Resumo:")
        print(f"  - Despesas: {count_desp}")
        print(f"  - Recebimentos: {count_recb}")
        print(f"  - Banco de dados: {os.path.abspath(db_path)}")
        return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Migrar dados do Excel para banco de dados SQLite'
    )
    parser.add_argument(
        '--xlsx',
        default='PLANILHA_ESTRATEGICA_MARCOS_2026.xlsx',
        help='Caminho da planilha Excel'
    )
    parser.add_argument(
        '--db',
        default='marcos.db',
        help='Caminho do arquivo SQLite'
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Output verboso'
    )
    
    args = parser.parse_args()
    
    success = migrate(args.xlsx, args.db, args.verbose)
    sys.exit(0 if success else 1)
